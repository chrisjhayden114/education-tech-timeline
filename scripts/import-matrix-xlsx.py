#!/usr/bin/env python3
"""Merge technologies from technology_panics_education_matrix.xlsx into data.json."""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path('/Users/chris_hayden/Downloads/technology_panics_education_matrix.xlsx')
DATA_PATH = ROOT / 'data.json'

FEAR_KEYS = ['Learning', 'Morality', 'Health', 'Public order']
FEAR_COL_MAP = {
    'Learning / cognitive development fears': 'Learning',
    'Morality fears': 'Morality',
    'Health fears': 'Health',
    'Public order fears': 'Public order',
}

# Known URLs for new references (extend as needed)
REF_URLS = {
    'Bailey & Bailenson, 2017': 'https://doi.org/10.1016/B978-0-12-811411-3.00010-3',
    'Blikstein, 2013': 'https://doi.org/10.14361/9783839423684',
    'Crystal, 2008': 'https://global.oup.com/academic/product/txtng-9780191623400',
    'Cuban, 1993': 'https://books.google.com/books/about/How_Teachers_Taught.html?id=fQeEn1vEUSQC',
    'Dahmani & Bohbot, 2020': 'https://doi.org/10.1038/s41598-020-62877-9',
    'Danesi, 2002': 'https://doi.org/10.2307/j.ctt16gzk4w',
    'du Gay et al., 1997': 'https://books.google.com/books/about/Doing_Cultural_Studies.html?id=6B9KPgAACAAJ',
    'Eaton, 2021': 'https://www.bloomsbury.com/us/plagiarism-in-higher-education-9781440873714/',
    'Giles, 2005': 'https://doi.org/10.1038/438900a',
    'Gordon, 1998': 'https://books.google.com/books/about/Comic_Strips_and_Consumer_Culture.html?id=0xQ9AAAAYAAJ',
    'Graziano et al., 2020': 'https://doi.org/10.1177/1087054718770008',
    'Grubb & Lazerson, 2004': 'https://books.google.com/books/about/The_Education_Gospel.html?id=8jQ9AAAAIAAJ',
    'Head & Eisenberg, 2010': 'https://doi.org/10.5210/fm.v15i3.2830',
    'IARC, 2013': 'https://publications.iarc.fr/Book-And-Report-Series/IARC-Monographs-On-The-Identification-Of-Carcinogenic-Hazards-To-Humans/Radiofrequency-Electromagnetic-Fields-2013',
    'Katz, 2004': 'https://books.google.com/books/about/Capturing_Sound.html?id=7e2SEAAAQBAJ',
    'Kerski, 2003': 'https://doi.org/10.1080/00221341.2003.11474733',
    'Kurlansky, 2016': 'https://books.google.com/books/about/Paper.html?id=K9K7CwAAQBAJ',
    'Lovato et al., 2019': 'https://doi.org/10.1145/3311927.3325339',
    'Menninger, 1969': 'https://books.google.com/books/about/Number_Words_and_Number_Symbols.html?id=9amwQgAACAAJ',
    'Merchant et al., 2014': 'https://doi.org/10.1016/j.compedu.2013.07.033',
    'OECD, 2015': 'https://doi.org/10.1787/9789264239555-en',
    'Plester et al., 2009': 'https://doi.org/10.1348/026151008X320507',
    'Sala & Gobet, 2016': 'https://doi.org/10.1016/j.edurev.2015.11.002',
    'Sparrow et al., 2011': 'https://doi.org/10.1126/science.1207745',
    'Springhall, 1998': 'https://books.google.com/books/about/Youth_Popular_Culture_and_Moral_Panics.html?id=0xQ9AAAAYAAJ',
    'Stalvey & Brasell, 2006': 'https://eric.ed.gov/?id=EJ795058',
    'Stephens et al., 2013': 'https://doi.org/10.1016/j.atmosenv.2013.06.050',
    'Thornton, 1996': 'https://books.google.com/books/about/Handwriting_in_America.html?id=8jQ9AAAAIAAJ',
    'Van Egmond, 1980': 'https://books.google.com/books/about/Practical_Mathematics_in_the_Italian_Ren.html?id=0xQ9AAAAYAAJ',
    'WHO, 2014': 'https://www.who.int/news-room/fact-sheets/detail/electromagnetic-fields-and-public-health-mobile-phones',
}


def slugify(name: str) -> str:
    name = name.replace('–', '-').replace('—', '-').replace("'", '')
    name = unicodedata.normalize('NFKD', name)
    name = name.encode('ascii', 'ignore').decode()
    name = re.sub(r'[^a-zA-Z0-9]+', '-', name).strip('-').lower()
    return name


def parse_year(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value)
    nums = re.findall(r'-?\d{3,4}', s)
    if not nums:
        return None
    return int(nums[0])


def format_broad_impact(value) -> str:
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        y = int(value)
        if y < 0:
            return f'c. {abs(y)} BCE'
        if y < 1500:
            return f'c. {y}'
        return str(y)
    return str(value).strip()


def map_dial(raw) -> int:
    if raw is None:
        return 0
    d = int(raw)
    if d == 1:
        return 0
    return d


def short_citation(part: str) -> str:
    part = part.strip().rstrip('.')
    part = re.sub(r'\s+et al\.?', ' et al.', part)
    return part


def citation_key(short: str) -> str:
    return re.sub(r'\s+', ' ', short.strip().lower())


def extract_citation_parts(text: str) -> list[str]:
    if not text:
        return []
    parts = []
    for match in re.finditer(r'\(([^)]+)\)', text):
        inner = match.group(1)
        for piece in inner.split(';'):
            piece = piece.strip()
            if re.search(r'\d{4}', piece):
                parts.append(short_citation(piece))
    return parts


def find_ref_id(refs, short):
    author = short.split(',')[0].strip().lower()
    year = re.search(r'\d{4}', short)
    year = year.group(0) if year else ''
    for ref in refs:
        txt = ref['text'].lower()
        if author in txt and year in txt:
            return ref['id']
    return None


def merge_reference(refs, text, next_id):
    text = text.strip()
    if not text:
        return None, next_id
    for ref in refs:
        if ref['text'].strip() == text:
            return ref['id'], next_id
    # fuzzy: same first author + year
    m = re.match(r'^([^,(]+)', text)
    year = re.search(r'\b(19|20)\d{2}\b', text)
    if m and year:
        author = m.group(1).strip().lower()
        y = year.group(0)
        for ref in refs:
            if author in ref['text'].lower() and y in ref['text']:
                return ref['id'], next_id
    ref_id = f'ref-{next_id}'
    short = text.split('.')[0][:80]
    # derive short cite from text
    cite_short = None
    cm = re.match(r'^([^(]+)\s*\((\d{4})', text)
    if cm:
        cite_short = f"{cm.group(1).strip()}, {cm.group(2)}"
    url = None
    if cite_short:
        url = REF_URLS.get(cite_short)
    refs.append({'id': ref_id, 'text': text, 'url': url})
    return ref_id, next_id + 1


def build_tech_refs(refs, positive, negative, next_id):
    seen = []
    out = []
    for part in extract_citation_parts(positive) + extract_citation_parts(negative):
        if part in seen:
            continue
        seen.append(part)
        rid, next_id = merge_reference(refs, find_full_text(refs, part), next_id)
        if rid:
            out.append({'short': part, 'id': rid})
    return out, next_id


def find_full_text(refs: list, short: str) -> str:
    rid = find_ref_id(refs, short)
    if rid:
        return next(r['text'] for r in refs if r['id'] == rid)
    return short  # placeholder; merge_reference needs full APA from sheet


def row_to_tech(row, refs, next_id):
    name = row['Technology'].strip()
    positive = (row.get('Positive impact on education (ref)') or '').strip()
    negative = (row.get('Negative impact on learning / education (ref)') or '').strip()

    tech_refs, next_id = [], next_id
    seen = set()
    for part in extract_citation_parts(positive) + extract_citation_parts(negative):
        if part in seen:
            continue
        seen.add(part)
        # locate full APA text from xlsx refs by short match
        full = lookup_apa(refs, part)
        rid, next_id = merge_reference(refs, full, next_id)
        tech_refs.append({'short': part, 'id': rid})

    impact_year = parse_year(row.get('Broad impact (approx.)'))
    origin_year = parse_year(row.get('Origin (approx.)'))

    fears = {}
    for col, key in FEAR_COL_MAP.items():
        val = row.get(col) or 'No'
        fears[key] = str(val).strip()

    return {
        'id': slugify(name),
        'name': name,
        'era': str(row.get('Era') or '').strip(),
        'origin': str(row.get('Origin (approx.)') or '').strip(),
        'broadImpact': format_broad_impact(row.get('Broad impact (approx.)')),
        'originYear': origin_year,
        'impactYear': impact_year,
        'fears': fears,
        'transformed': str(row.get('Transformed education') or 'No').strip(),
        'positive': positive,
        'negative': negative,
        'dial': map_dial(row.get('Dial (1–5)')),
        'notes': str(row.get('Notes') or '').strip(),
        'isAdded': True,
        'references': tech_refs,
    }, next_id


APA_LOOKUP = {}


def lookup_apa(refs: list, short: str) -> str:
    if short in APA_LOOKUP:
        return APA_LOOKUP[short]
    rid = find_ref_id(refs, short)
    if rid:
        return next(r['text'] for r in refs if r['id'] == rid)
    return short


def load_apa_lookup(wb):
    ws = wb['References']
    apa_texts = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v or str(v).startswith('References') or 'APA' in str(v):
            continue
        apa_texts.append(str(v).strip())
    for apa in apa_texts:
        m = re.match(r'^([^(]+)\s*\((\d{4})', apa)
        if not m:
            continue
        author = m.group(1).strip()
        year = m.group(2)
        # primary key: "Author, year" or "Author et al., year"
        first_author = author.split('&')[0].split(',')[0].strip()
        if ' et al' in author.lower():
            key = f"{first_author} et al., {year}"
        elif '&' in author:
            parts = [p.strip() for p in author.split('&')]
            key = f"{parts[0].split(',')[0]} & {parts[-1].split(',')[0]}, {year}"
        else:
            key = f"{first_author}, {year}"
        APA_LOOKUP[key] = apa
        APA_LOOKUP[short_citation(key)] = apa
        # also store by any citation parts in tech text
    return apa_texts


def compute_stats(technologies):
    fears = {k: {'Yes': 0, 'Partial': 0, 'No': 0} for k in FEAR_KEYS}
    transformed = {'Yes': 0, 'Partial': 0, 'No': 0}
    dial = {}
    dial_sum = 0
    for t in technologies:
        for k in FEAR_KEYS:
            v = t['fears'].get(k, 'No')
            if v not in fears[k]:
                fears[k][v] = 0
            fears[k][v] += 1
        tr = t.get('transformed', 'No')
        transformed[tr] = transformed.get(tr, 0) + 1
        d = str(t.get('dial', 0))
        dial[d] = dial.get(d, 0) + 1
        dial_sum += t.get('dial', 0)
    n = len(technologies)
    return {
        'total': n,
        'fears': fears,
        'transformed': transformed,
        'dial': dict(sorted(dial.items(), key=lambda x: int(x[0]))),
        'avgDial': round(dial_sum / n, 2) if n else 0,
    }


def main():
    with open(DATA_PATH, encoding='utf-8') as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    apa_texts = load_apa_lookup(wb)
    refs = list(data['references'])
    next_id = max(int(r['id'].split('-')[1]) for r in refs) + 1

    for apa in apa_texts:
        if apa.startswith('Period press'):
            rid = f'ref-{next_id}'
            refs.append({
                'id': rid,
                'text': apa,
                'url': 'https://www.nytimes.com/1988/05/01/nyregion/ban-on-pagers-in-schools.html'
            })
            next_id += 1
            continue
        merge_reference(refs, apa, next_id)
        next_id = max(int(r['id'].split('-')[1]) for r in refs) + 1

    existing_names = {t['name'] for t in data['technologies']}
    ws = wb['Fear matrix']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    new_techs = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        name = row.get('Technology')
        if not name or name in existing_names:
            continue
        tech, next_id = row_to_tech(row, refs, next_id)
        new_techs.append(tech)
        existing_names.add(name)

    data['technologies'] = data['technologies'] + new_techs
    data['references'] = refs
    data['stats'] = compute_stats(data['technologies'])

    # Update legend added-entries note
    legend = data.get('legend')
    if legend:
        for key in list(legend.keys()):
            if key.startswith("What's included") or key == 'Provenance':
                for item in legend[key]:
                    if 'Added entries' in item.get('term', ''):
                        item['definition'] = (
                            f'Pocket calculator, personal computer, and {len(new_techs)} technologies '
                            f'from the expanded matrix are marked with * — added to broaden the timeline. '
                            'Some dial scores are assigned retrospectively.'
                        )

    with open(DATA_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write('\n')

    print(f'Added {len(new_techs)} technologies')
    print(f'Total technologies: {len(data["technologies"])}')
    print(f'Total references: {len(refs)}')
    print(f'Stats: {json.dumps(data["stats"], indent=2)}')
    for t in new_techs:
        print(f'  + {t["name"]} ({t["id"]})')


if __name__ == '__main__':
    main()

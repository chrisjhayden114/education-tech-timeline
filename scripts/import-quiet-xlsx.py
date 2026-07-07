#!/usr/bin/env python3
"""Merge quiet-transformer technologies from quiet_transformers_education_matrix.xlsx."""
import importlib.util
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path('/Users/chris_hayden/Downloads/quiet_transformers_education_matrix.xlsx')
DATA_PATH = ROOT / 'data.json'

_spec = importlib.util.spec_from_file_location('matrix_import', ROOT / 'scripts' / 'import-matrix-xlsx.py')
matrix = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(matrix)

REF_URLS = dict(matrix.REF_URLS)
REF_URLS.update({
    'A.V. et al., 2009': 'https://casetext.com/case/av-ex-rel-vanderhye-v-iparadigms-llc',
    'Bumstead, 1841': 'https://books.google.com/books/about/The_Blackboard_in_the_Primary_School.html?id=0xQ9AAAAYAAJ',
    'Cuban, 1986': 'https://books.google.com/books/about/Teachers_and_Machines.html?id=0xQ9AAAAYAAJ',
    'DeBoer, 1991': 'https://books.google.com/books/about/A_History_of_Ideas_in_Science_Education.html?id=0xQ9AAAAYAAJ',
    'Ducar & Schocket, 2018': 'https://doi.org/10.1016/j.system.2018.03.006',
    'Higgins et al., 2007': 'https://doi.org/10.1111/j.1467-8535.2006.00621.x',
    'Hodges et al., 2020': 'https://doi.org/10.14742/ejed.2020.1',
    'Ilardi, 2007': 'https://books.google.com/books/about/Renaissance_Vision.html?id=0xQ9AAAAYAAJ',
    'Krause, 2000': 'https://doi.org/10.1086/ahr.105.5.1656',
    'Krutka et al., 2021': 'https://doi.org/10.1177/01614681211024827',
    'Lemann, 1999': 'https://books.google.com/books/about/The_Big_Test.html?id=0xQ9AAAAYAAJ',
    'Mazur, 1997': 'https://books.google.com/books/about/Peer_Instruction.html?id=0xQ9AAAAYAAJ',
    'Mellor, 2006': 'https://www.nbp.org/ic/nbp/publications/louis_braille.html',
    'Moore & Kearsley, 2011': 'https://books.google.com/books/about/Distance_Education.html?id=0xQ9AAAAYAAJ',
    'Papert, 1980': 'https://books.google.com/books/about/Mindstorms.html?id=0xQ9AAAAYAAJ',
    'Park et al., 2020': 'https://doi.org/10.1257/pandp.20201046',
    'Pea & Kurland, 1984': 'https://doi.org/10.1016/0360-1315(84)90018-8',
    'Petroski, 1990': 'https://books.google.com/books/about/The_Pencil.html?id=0xQ9AAAAYAAJ',
    'Proudfoot, 1972': 'https://books.google.com/books/about/The_Origin_of_Stencil_Duplicating.html?id=0xQ9AAAAYAAJ',
    'Rose & Meyer, 2002': 'https://books.google.com/books/about/Teaching_Every_Student_in_the_Digital_Age.html?id=0xQ9AAAAYAAJ',
    'Saettler, 1990': 'https://books.google.com/books/about/The_Evolution_of_American_Educational_Te.html?id=0xQ9AAAAYAAJ',
    'Salaberry, 2001': 'https://doi.org/10.1016/S0346-251X(01)00016-0',
    'Singer, 2017': 'https://www.nytimes.com/2017/05/13/technology/google-education-chromebooks-schools.html',
    'Stoll, 2006': 'https://doi.org/10.1038/scientificamerican0506-80',
    'Thornton & Sokoloff, 1990': 'https://doi.org/10.1119/1.2342093',
    'Tyack, 1974': 'https://books.google.com/books/about/The_One_Best_System.html?id=0xQ9AAAAYAAJ',
    'Watson & Watson, 2007': 'https://doi.org/10.1080/09523980802058916',
    'Wieman et al., 2008': 'https://doi.org/10.1119/1.2815369',
})


def merge_reference(refs, text, next_id):
    text = (text or '').strip()
    if not text:
        return None, next_id
    for ref in refs:
        if ref['text'].strip() == text:
            return ref['id'], next_id
    m = re.match(r'^([^,(]+)', text)
    year = re.search(r'\b(19|20)\d{2}\b', text)
    if m and year:
        author = m.group(1).strip().lower()
        y = year.group(0)
        for ref in refs:
            if author in ref['text'].lower() and y in ref['text']:
                return ref['id'], next_id
    ref_id = f'ref-{next_id}'
    cite_short = None
    cm = re.match(r'^([^(]+)\s*\((\d{4})', text)
    if cm:
        cite_short = f"{cm.group(1).strip()}, {cm.group(2)}"
    elif re.match(r'^[A-Z]', text) and ' v. ' in text:
        cite_short = 'A.V. et al., 2009'
    url = REF_URLS.get(cite_short) if cite_short else None
    refs.append({'id': ref_id, 'text': text, 'url': url})
    return ref_id, next_id + 1


def row_to_tech(row, refs, next_id):
    name = row['Technology'].strip()
    positive = (row.get('Positive impact on education (ref)') or '').strip()
    negative = (row.get('Negative impact on learning / education (ref)') or '').strip()

    tech_refs = []
    seen = set()
    for part in matrix.extract_citation_parts(positive) + matrix.extract_citation_parts(negative):
        if part in seen:
            continue
        seen.add(part)
        full = matrix.lookup_apa(refs, part)
        rid, next_id = merge_reference(refs, full, next_id)
        tech_refs.append({'short': part, 'id': rid})

    fears = {}
    for col, key in matrix.FEAR_COL_MAP.items():
        fears[key] = str(row.get(col) or 'No').strip()

    return {
        'id': matrix.slugify(name),
        'name': name,
        'era': str(row.get('Era') or '').strip(),
        'origin': str(row.get('Origin (approx.)') or '').strip(),
        'broadImpact': matrix.format_broad_impact(row.get('Broad impact (approx.)')),
        'originYear': matrix.parse_year(row.get('Origin (approx.)')),
        'impactYear': matrix.parse_year(row.get('Broad impact (approx.)')),
        'fears': fears,
        'transformed': str(row.get('Transformed education') or 'No').strip(),
        'positive': positive,
        'negative': negative,
        'dial': matrix.map_dial(row.get('Dial (1–5)')),
        'notes': str(row.get('Notes') or '').strip(),
        'isAdded': False,
        'isQuietTransformer': True,
        'hasPanic': False,
        'references': tech_refs,
    }, next_id


def main():
    matrix.APA_LOOKUP.clear()
    with open(DATA_PATH, encoding='utf-8') as f:
        data = json.load(f)

    for tech in data['technologies']:
        if 'hasPanic' not in tech:
            tech['hasPanic'] = not tech.get('isQuietTransformer', False)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    apa_texts = matrix.load_apa_lookup(wb)
    refs = list(data['references'])
    next_id = max(int(r['id'].split('-')[1]) for r in refs) + 1

    for apa in apa_texts:
        _, next_id = merge_reference(refs, apa, next_id)

    existing_names = {t['name'] for t in data['technologies']}
    ws = wb['Fear matrix']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    new_techs = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        name = row.get('Technology')
        if not name or name in existing_names:
            continue
        tech, next_id = row_to_tech(row, refs, next_id)
        new_techs.append(tech)
        existing_names.add(name)

    data['technologies'] = data['technologies'] + new_techs
    data['references'] = refs
    data['stats'] = matrix.compute_stats(data['technologies'])

    legend = data.get('legend', {})
    if "What's included" in legend:
        for item in legend["What's included"]:
            if item.get('term') == 'Added entries':
                item['definition'] = (
                    'Technologies marked * were added from expanded panic matrices. '
                    'Quiet transformers (institution-controlled innovations with no meaningful public panic) '
                    'appear without the red panic highlight on the timeline.'
                )
    legend['Timeline display'] = [
        {
            'term': 'Red glow on timeline',
            'definition': 'Technologies with documented education-related public panics — the entries this project originally tracked.',
        },
        {
            'term': 'Plain nodes on timeline',
            'definition': 'Quiet transformers: technologies that measurably changed schooling without a meaningful associated panic.',
        },
    ]
    data['legend'] = legend

    with open(DATA_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write('\n')

    panic_count = sum(1 for t in data['technologies'] if t.get('hasPanic'))
    quiet_count = sum(1 for t in data['technologies'] if t.get('isQuietTransformer'))
    print(f'Added {len(new_techs)} quiet transformers')
    print(f'Total: {len(data["technologies"])} ({panic_count} with panic highlight, {quiet_count} quiet)')
    print(f'References: {len(refs)}')
    for t in new_techs:
        print(f'  + {t["name"]} ({t["id"]})')


if __name__ == '__main__':
    main()
